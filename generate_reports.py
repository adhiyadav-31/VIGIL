import os
import openpyxl
from fpdf import FPDF

# The server is running from TKR_404-main, so downloads should be in frontend/downloads or just root.
# The server is serving from backend/frontend perhaps, let's put it in frontend/reports
out_dir = r"c:\Users\GUDA ADHI YADAV\Downloads\TKR_404-main\app\frontend\reports"
os.makedirs(out_dir, exist_ok=True)

reports = ["Monthly Report", "Quarterly Report", "Risk Report", "Growth Report"]

for r in reports:
    fname = r.replace(" ", "_")
    
    # Generate Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = r
    ws['A1'] = f"VigilAI {r}"
    ws['A2'] = "Generated for Meridian Textiles"
    ws['A4'] = "Metric"
    ws['B4'] = "Value"
    ws['A5'] = "Revenue"
    ws['B5'] = "₹ 1,80,000"
    ws['A6'] = "Expenses"
    ws['B6'] = "₹ 1,40,000"
    ws['A7'] = "Net Profit"
    ws['B7'] = "₹ 40,000"
    wb.save(os.path.join(out_dir, f"{fname}.xlsx"))
    
    # Generate PDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=15)
    pdf.cell(200, 10, txt=f"VigilAI {r}", ln=True, align='C')
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt="Generated for Meridian Textiles", ln=True, align='C')
    pdf.cell(200, 10, txt="", ln=True, align='C')
    pdf.cell(200, 10, txt="Revenue: INR 1,80,000", ln=True, align='L')
    pdf.cell(200, 10, txt="Expenses: INR 1,40,000", ln=True, align='L')
    pdf.cell(200, 10, txt="Net Profit: INR 40,000", ln=True, align='L')
    pdf.output(os.path.join(out_dir, f"{fname}.pdf"))
